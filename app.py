# app.py
from flask import Flask, render_template, request, jsonify
import numpy as np
import serial
import time
import threading
from collections import deque
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
from openpyxl.chart import (
    LineChart,
    Reference
)

SERIAL_PORT = 'COM5'
BAUD_RATE = 57600

latest_data = {"Ra": 0.0, "Rb": 0.0, "P": 0.0, "load_pos": 0.0}
data_lock = threading.Lock()

# Global Software Tare Baselines (Tracking hardware drift on the fly)
offset_r1 = 0.0
offset_r2 = 0.0

# UI Warning State
system_busy = True
last_stable_load = 0.0
load_change_threshold = 0.15   # kN
stable_counter = 0

last_logged_snapshot = None

# State tracking mechanics
has_hot_tared = False
first_load_detected = False  # Prevents startup 0-point from triggering hot-tare prematurely

app = Flask(__name__)

# ========================= KALMAN FILTER =========================
class KalmanFilter1D:
    def __init__(self, process_variance=0.005, measurement_variance=1.5, initial_value=0.0):
        self.process_variance = process_variance      
        self.measurement_variance = measurement_variance  
        self.estimate = initial_value
        self.error_estimate = 1.0

    def update(self, measurement):
        predicted_estimate = self.estimate
        predicted_error = self.error_estimate + self.process_variance
        kalman_gain = predicted_error / (predicted_error + self.measurement_variance)
        
        self.estimate = predicted_estimate + kalman_gain * (measurement - predicted_estimate)
        self.error_estimate = (1 - kalman_gain) * predicted_error
        return self.estimate

# Kalman Filters
kalman_ra = KalmanFilter1D()
kalman_rb = KalmanFilter1D()

# ========================= MOVING AVERAGE =========================
MA_WINDOW = 1   

ma_buffer_ra = deque(maxlen=MA_WINDOW)
ma_buffer_rb = deque(maxlen=MA_WINDOW)

def moving_average(buffer, new_value):
    buffer.append(new_value)
    if len(buffer) == 0:
        return new_value
    return sum(buffer) / len(buffer)

# ================================================================


# =====================================================
# TEST SESSION LOGGER
# =====================================================

test_history = deque(maxlen=10000)
last_logged_time = 0
LOG_INTERVAL = 1.0   # seconds

def read_serial():
    global latest_data, offset_r1, offset_r2
    global has_hot_tared, first_load_detected
    global system_busy, last_stable_load, stable_counter
    global last_logged_time
    global last_logged_snapshot
    try:
        ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=1.5)
        time.sleep(2.5) # Wait for Arduino reboot cycle
        print(f"✅ Connected to Arduino on {SERIAL_PORT}")

        # ----------------------------------------------------------------------
        # AUTOMATED INITIAL HARDWARE TARE HANDSHAKE
        # ----------------------------------------------------------------------
        print("\n" + "="*70)
        print("🚨 [PRE-INITIALIZATION] AUTOMATED SCALE TARE STARTED")
        print("👉 SYSTEM ASSUMES HANGER HOOK IS ENGAGED ON THE BEAM ASSEMBLY NOW!")
        print("="*70)
        
        for count in range(5, 0, -1):
            print(f"⏳ Sending automated tare in {count} seconds...")
            time.sleep(1.0)
            
        print("\n⚙️ Sending electronic tare character ('t') to Arduino...")
        ser.reset_input_buffer()  
        ser.write(b't')
        ser.flush()

        tared_confirmed = False
        start_wait = time.time()
        
        system_busy = True
        print("⏳ Waiting for hardware load cell tare confirmation...")
        while not tared_confirmed and (time.time() - start_wait < 6.0):
            if ser.in_waiting > 0:
                confirmation_line = ser.readline().decode('utf-8', errors='ignore').strip()
                print(f"   [MCU ECHO]: {confirmation_line}")
                if "Both scales tared." in confirmation_line:
                    tared_confirmed = True
            time.sleep(0.05)

        if tared_confirmed:
            print("✨ Hardware confirmation received successfully.")
        else:
            print("⚠️ Tare confirmation timeout reached. Moving forward assuming hardware zero-point updated.")

        time.sleep(1.5)
        print("\n" + "="*70)
        print("✅ [TARE COMPLETE - LIVE SYSTEM MONITOR RUNNING]")
        system_busy = False
        print("👉 FIRST RUN RUNNING: HOOK EXCLUDED SUCCESSFULLY BY HARDWARE PROFILE.")
        print("="*70 + "\n")
        
        ser.reset_input_buffer() 
        # ----------------------------------------------------------------------

        # Tracker to prevent continuous spam printing during a single auto-tare event
        is_tared_empty = False 

        while True:
            if ser.in_waiting > 0:
                line = ser.readline().decode('utf-8', errors='ignore').strip()
                
                if line and ',' in line and not line.startswith("Both"):
                    try:
                        # Raw values streamed straight from the Arduino scale (in GRAMS)
                        raw_r1, raw_r2 = map(float, line.split(','))   
                        
                        # Apply live software offsets to handle structural drift 
                        r1 = raw_r1 - offset_r1
                        r2 = raw_r2 - offset_r2

                        # Convert from grams to kN
                        r1_kn = (r1 / 1000.0) * 9.81
                        r2_kn = (r2 / 1000.0) * 9.81

                        # Filters (Optimized for faster convergence)
                        filtered_ra = kalman_ra.update(r1_kn)
                        filtered_rb = kalman_rb.update(r2_kn)

                        smoothed_ra = moving_average(ma_buffer_ra, filtered_ra)
                        smoothed_rb = moving_average(ma_buffer_rb, filtered_rb)

                        # Calculate active sum for the tare trigger threshold check
                        current_p = smoothed_ra + smoothed_rb
                        
                        # Detect sudden load changes
                        if abs(current_p - last_stable_load) > load_change_threshold:
                            system_busy = True
                            stable_counter = 0

                        # Arm the trigger mechanism once a true weight is detected (> 50 grams payload)
                        if not first_load_detected and current_p > ((50.0 / 1000.0) * 9.81):
                            first_load_detected = True
                            print("🚀 [SYSTEM INSIGHT] First true test load detected. Tare triggers armed.")

                        # 🔄 AUTOMATIC SOFTWARE HOT-TARE TRIGGER 🔄
                        # Only allow it to trip if a load was previously verified on the structure
                        if current_p < 0.0001 and first_load_detected:
                            system_busy = True 
                            if not is_tared_empty:
                                # Snapshot bare-beam mechanical state layout (hook is off)
                                offset_r1 = raw_r1
                                offset_r2 = raw_r2
                                
                                # Flush filters back to dead absolute zero
                                kalman_ra.estimate = 0.0
                                kalman_rb.estimate = 0.0
                                ma_buffer_ra.clear()
                                ma_buffer_rb.clear()
                                
                                print("\n" + "-"*50)
                                print(f"🔄 [MID-TEST AUTO-TARE DETECTED] Load assemblies removed from rig.")
                                print(f"   ↳ Hook mass (1572g) correction will apply on all future measurements.")
                                print("-"*50 + "\n")
                                
                                is_tared_empty = True
                                has_hot_tared = True  
                            
                            smoothed_ra = 0.0
                            smoothed_rb = 0.0
                            current_p = 0.0
                            if abs(current_p) < 0.0001:
                                system_busy = False
                        else:
                            is_tared_empty = False

                        # Hook correction application factor in kN (1572g)
                        HOOK_WEIGHT_KN = (1572.0 / 1000.0) * 9.81
                        
                        if system_busy:
                            if abs(current_p - last_stable_load) < 0.03:
                                stable_counter += 1
                            else:
                                stable_counter = 0

                            # About 1 second of stable data
                            if stable_counter > 40:
                                system_busy = False
                                last_stable_load = current_p

                        with data_lock:
                            if not has_hot_tared:
                                # Run #1: Leave reading untouched (Hook is already zeroed by hardware)
                                latest_data["Ra"] = round(smoothed_ra, 4)   
                                latest_data["Rb"] = round(smoothed_rb, 4)
                                latest_data["P"] = round(current_p, 4)
                            else:
                                # Run #2 onwards: Subtract the 1572g hook load from the live values
                                latest_data["P"] = round(max(current_p - HOOK_WEIGHT_KN, 0.0), 4)
                                
                                if latest_data["P"] > 0.001:
                                    total_physical = current_p if current_p > 0 else 1.0
                                    ratio_b = smoothed_rb / total_physical
                                    latest_data["Rb"] = round(latest_data["P"] * ratio_b, 4)
                                    latest_data["Ra"] = round(latest_data["P"] * (1.0 - ratio_b), 4)
                                else:
                                    latest_data["Rb"] = 0.0
                                    latest_data["Ra"] = 0.0
                                
                            # Position Engine Calculation
                            if latest_data["P"] > 0.01:   
                                L = 1.0
                                if has_hot_tared:
                                    computed_pos = (smoothed_rb / current_p) * L
                                else:
                                    computed_pos = (latest_data["Rb"] / latest_data["P"]) * L
                                latest_data["load_pos"] = round(min(max(computed_pos, 0.0), L), 4)
                            else:
                                latest_data["load_pos"] = 0.0
                        # =====================================================
                        # SESSION LOGGING
                        # =====================================================

                        current_time = time.time()

                        if not system_busy and (current_time - last_logged_time) >= LOG_INTERVAL:

                            max_sfd = max(
                                abs(latest_data["Ra"]),
                                abs(latest_data["Ra"] - latest_data["P"])
                            )

                            max_bmd = (
                                latest_data["Ra"]
                                * latest_data["load_pos"]
                            )

                            snapshot = (
                                round(latest_data["P"], 3),
                                round(latest_data["load_pos"], 3)
                            )

                            if snapshot != last_logged_snapshot:

                                test_history.append({
                                    "timestamp": datetime.now().strftime("%H:%M:%S"),
                                    "load": latest_data["P"],
                                    "position": latest_data["load_pos"],
                                    "Ra": latest_data["Ra"],
                                    "Rb": latest_data["Rb"],
                                    "max_sfd": max_sfd,
                                    "max_bmd": max_bmd
                                })

                                last_logged_snapshot = snapshot

                            last_logged_time = current_time

                    except Exception as e:
                        print(f"Parsing error: {e}")
                        
            time.sleep(0.02) 
    except Exception as e:
        print(f"Serial Error: {e}")

threading.Thread(target=read_serial, daemon=True).start()

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/data")
def data():
    try:
        L = float(request.args.get("L", 1.0))
        if L > 1.0: L = 1.0
        L = max(L, 0.1)   
    
        E = float(request.args.get("E", 210.0)) * 1e9
        I = float(request.args.get("I", 0.039852)) * 1e-6
        y = float(request.args.get("y", 15.0)) / 1000
        x_query = float(request.args.get("x", 0))

        with data_lock:
            Ra_corrected = latest_data["Ra"]
            Rb_corrected = latest_data["Rb"]
            P_corrected = latest_data["P"]
            load_pos = latest_data["load_pos"]

        load_pos = min(max(load_pos, 0.0), L)
        
        if P_corrected < 1e-3:  
            P_corrected = 0.0   
            load_pos = 0.0
            Ra_corrected = 0.0
            Rb_corrected = 0.0

        x = np.linspace(0, L, 400)
        
        V = np.where(x < load_pos, Ra_corrected, Ra_corrected - P_corrected)
        M = np.where(x < load_pos, Ra_corrected * x, Ra_corrected * x - P_corrected * (x - load_pos))

        delta = np.zeros_like(x)
        a = load_pos
        b = L - a
        for i, xi in enumerate(x):
            if xi <= a:
                delta[i] = (P_corrected * b * xi) / (6 * L * E * I) * (L**2 - b**2 - xi**2)
            else:
                delta[i] = (P_corrected * a * (L - xi)) / (6 * L * E * I) * (L**2 - a**2 - (L - xi)**2)

        if x_query < load_pos:
            Vx = Ra_corrected
            Mx = Ra_corrected * x_query
        else:
            Vx = Ra_corrected - P_corrected
            Mx = Ra_corrected * x_query - P_corrected * (x_query - load_pos)

        sigma = (Mx * y) / I if I > 0 else 0

        return jsonify({
            "P": float(round(P_corrected, 4)),
            "Ra": float(round(Ra_corrected, 4)),
            "Rb": float(round(Rb_corrected, 4)),
            "load_pos": float(load_pos),
            "L": float(L),
            "V": V.tolist(),
            "M": M.tolist(),
            "delta": delta.tolist(),
            "X": x.tolist(),
            "Vx": float(Vx),
            "Mx": float(Mx),
            "sigma": float(sigma),
            "busy": system_busy,
            "delta_x": float(delta[np.argmin(np.abs(x - x_query))])
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/export")
def export_data():

    try:

        L = float(request.args.get("L", 1.0))
        E = float(request.args.get("E", 210))
        I = float(request.args.get("I", 0.039852))
        y = float(request.args.get("y", 15))

        with data_lock:
            Ra = latest_data["Ra"]
            Rb = latest_data["Rb"]
            P = latest_data["P"]
            load_pos = latest_data["load_pos"]

        wb = Workbook()

        # ====================================================
        # SUMMARY SHEET
        # ====================================================

        ws = wb.active
        ws.title = "Test Information"
        
        ws["A1"] = "BEAMLAB PRO TEST REPORT"
        ws["A1"].font = Font(bold=True, size=18)

        ws["A3"] = "Team Name"
        ws["B3"] = "Abenezer / Simrgeta / Waktole / Israel / Tesfahun / Filimon"

        ws["A4"] = "Date"
        ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")

        ws["A5"] = "Specimen"
        ws["B5"] = "Simply Supported Beam"

        ws["A6"] = "Beam Length (m)"
        ws["B6"] = L

        title_fill = PatternFill(
            start_color="00AA88",
            end_color="00AA88",
            fill_type="solid"
        )

        title_font = Font(
            bold=True,
            color="FFFFFF",
            size=14
        )

        headers = [
            "Parameter",
            "Value"
        ]

        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = text
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal="center")

        rows = [
            ["Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Beam Length L (m)", L],
            ["Load Position (m)", load_pos],
            ["Applied Load P (N)", P],
            ["Reaction Ra (N)", Ra],
            ["Reaction Rb (N)", Rb],
            ["Elastic Modulus E (GPa)", E],
            ["Moment of Inertia I", I],
            ["Distance y (mm)", y]
        ]

        for r, values in enumerate(rows, start=2):
            ws.cell(row=r, column=1, value=values[0])
            ws.cell(row=r, column=2, value=values[1])

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25

        # ====================================================
        # SFD/BMD SHEET
        # ====================================================

        ws2 = wb.create_sheet("Beam Data")

        headers = [
            "Position (m)",
            "Shear Force (N)",
            "Bending Moment (Nm)"
        ]

        for c, text in enumerate(headers, start=1):
            cell = ws2.cell(row=1, column=c)
            cell.value = text
            cell.fill = title_fill
            cell.font = title_font

        x = np.linspace(0, L, 400)

        V = np.where(
            x < load_pos,
            Ra,
            Ra - P
        )

        M = np.where(
            x < load_pos,
            Ra * x,
            Ra * x - P * (x - load_pos)
        )

        for i in range(len(x)):
            ws2.append([
                float(x[i]),
                float(V[i]),
                float(M[i])
            ])

        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 20
        ws2.column_dimensions['C'].width = 20
        
        history_sheet = wb.create_sheet("Reading History")
        headers = [
            "Time",
            "Load (N)",
            "Position (m)",
            "Ra (N)",
            "Rb (N)",
            "Max SFD (N)",
            "Max BMD (Nm)"
        ]
        
        for c, h in enumerate(headers, start=1):
            history_sheet.cell(
                row=1,
                column=c,
                value=h
            )
            
        for row in test_history:

            history_sheet.append([
                row["timestamp"],
                row["load"],
                row["position"],
                row["Ra"],
                row["Rb"],
                row["max_sfd"],
                row["max_bmd"]
            ])
            
        sfd_sheet = wb.create_sheet("SFD Data")
        
        sfd_sheet.append([
            "Position (m)",
            "Shear Force (N)"
        ])

        for i in range(len(x)):
            sfd_sheet.append([
                float(x[i]),
                float(V[i])
            ])
        
        bmd_sheet = wb.create_sheet("BMD Data")
        bmd_sheet.append([
            "Position (m)",
            "Moment (Nm)"
        ])

        for i in range(len(x)):
            bmd_sheet.append([
                float(x[i]),
                float(M[i])
            ])
            
        def_sheet = wb.create_sheet("Deflection Data")

        def_sheet.append([
            "Position (m)",
            "Deflection (m)"
        ])
        
        # ====================================================
        # DEFLECTION CALCULATION
        # ====================================================

        delta = np.zeros_like(x)

        E_pa = E * 1e9          # Convert GPa -> Pa
        I_m4 = I * 1e-6         # Same conversion used elsewhere

        a = load_pos
        b = L - a

        for i, xi in enumerate(x):
            if xi <= a:
                delta[i] = (
                    (P * b * xi)
                    / (6 * L * E_pa * I_m4)
                    * (L**2 - b**2 - xi**2)
                )
            else:
                delta[i] = (
                    (P * a * (L - xi))
                    / (6 * L * E_pa * I_m4)
                    * (L**2 - a**2 - (L - xi)**2)
                )

        for i in range(len(x)):
            def_sheet.append([
                float(x[i]),
                float(delta[i])
            ])
        
        chart_sheet = wb.create_sheet("Charts")
        
        chart1 = LineChart()

        chart1.title = "Shear Force Diagram"

        data = Reference(
            sfd_sheet,
            min_col=2,
            min_row=1,
            max_row=len(x)+1
        )

        cats = Reference(
            sfd_sheet,
            min_col=1,
            min_row=2,
            max_row=len(x)+1
        )

        chart1.add_data(
            data,
            titles_from_data=True
        )

        chart1.set_categories(cats)

        chart_sheet.add_chart(
            chart1,
            "A1"
        )
        
        chart2 = LineChart()

        chart2.title = "Bending Moment Diagram"

        data = Reference(
            bmd_sheet,
            min_col=2,
            min_row=1,
            max_row=len(x)+1
        )

        cats = Reference(
            bmd_sheet,
            min_col=1,
            min_row=2,
            max_row=len(x)+1
        )

        chart2.add_data(
            data,
            titles_from_data=True
        )

        chart2.set_categories(cats)

        chart_sheet.add_chart(
            chart2,
            "A18"
        )
        
        chart3 = LineChart()

        chart3.title = "Deflection Curve"

        data = Reference(
            def_sheet,
            min_col=2,
            min_row=1,
            max_row=len(x)+1
        )

        cats = Reference(
            def_sheet,
            min_col=1,
            min_row=2,
            max_row=len(x)+1
        )

        chart3.add_data(
            data,
            titles_from_data=True
        )

        chart3.set_categories(cats)

        chart_sheet.add_chart(
            chart3,
            "J1"
        )
        
        stats_sheet = wb.create_sheet("Statistics")
        if test_history:

            loads = [r["load"] for r in test_history]

            stats_sheet.append(["Maximum Load", max(loads)])
            stats_sheet.append(["Minimum Load", min(loads)])
            stats_sheet.append(["Average Load", np.mean(loads)])
            stats_sheet.append(["Number of Samples", len(loads)])
        
        # ====================================================
        # SAVE TO MEMORY
        # ====================================================

        excel_file = BytesIO()

        wb.save(excel_file)

        excel_file.seek(0)

        filename = (
            f"BEAMLAB_TEST_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)})


@app.route("/clear_session")
def clear_session():

    global test_history

    test_history.clear()

    return jsonify({
        "status": "Session cleared"
    })


if __name__ == "__main__":
    print("🚀 Starting Beam Lab Server - Auto Hot-Tare Sequence Armed")
    app.run(debug=True, port=5000, use_reloader=False)